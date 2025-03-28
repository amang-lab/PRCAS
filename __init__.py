import math
import os
from collections import defaultdict
import cv2
import numpy as np
from openpyxl.styles import PatternFill
from ultralytics import YOLO
import lane_detection as ld
import matplotlib.path as mplPath
import openpyxl
from openpyxl import load_workbook
import pandas as pd

"""
This is the file to extra data from video for training
"""

# Function to draw bounding boxes without ID and confidence
def draw_boxes(image, boxes):
    for box in boxes:
        x1, y1, x2, y2 = map(int, box[:4])
        cv2.rectangle(image, (x1, y1), (x2, y2), (0, 0, 255), 2)


def draw_boxes_selected(image, boxes, class_ids, selected_classes):
    for box, class_id in zip(boxes, class_ids):
        if class_id in selected_classes:
            x1, y1, x2, y2 = map(int, box[:4])
            cv2.rectangle(image, (x1, y1), (x2, y2), (0, 0, 255), 2)


def find_points_dividing_line(A, B, ratio):
    (x1, y1) = A
    (x2, y2) = B

    mid_x = (x1 + (ratio * x2)) / (1 + ratio)
    mid_y = (y1 + (ratio * y2)) / (1 + ratio)
    return (int(mid_x), int(mid_y))


def direction(A, B):
    (x1, y1) = A
    (x2, y2) = B
    return x2 - x1


def is_inside(point, polygon):
    return polygon.contains_point(point)


def is_point_in_trapezoid(p, a, b, c, d):
    def cross_product(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    def is_point_in_triangle(p, a, b, c):
        cp1 = cross_product(p, a, b)
        cp2 = cross_product(p, b, c)
        cp3 = cross_product(p, c, a)
        return (cp1 >= 0 and cp2 >= 0 and cp3 >= 0) or (cp1 <= 0 and cp2 <= 0 and cp3 <= 0)

    # Divide trapezoid into two triangles
    return is_point_in_triangle(p, a, b, d) or is_point_in_triangle(p, b, c, d)


def slope(A, B):
    (x1, y1) = A
    (x2, y2) = B
    if x2 - x1 != 0:
        return (y2 - y1) / (x2 - x1)
    else:
        return 0


def angleofxaxis(slope):
    angle_radians = math.atan(slope)
    angle_degrees = math.degrees(angle_radians)
    return angle_degrees


def angle_by_slop(m1, m2):
    # Calculate the tangent of the angle
    tan_theta = abs((m2 - m1) / (1 + m1 * m2))
    # Calculate the angle in radians
    theta_radians = math.atan(tan_theta)
    # Convert the angle to degrees
    theta_degrees = math.degrees(theta_radians)
    return theta_degrees


def features_from_base_point(one_zone, start_point, end_point):
    left_slope = slope(one_zone.vertices[0], one_zone.vertices[3])
    right_slope = slope(one_zone.vertices[1], one_zone.vertices[2])
    line_slope = slope(start_point, end_point)

    angle_of_moving = angleofxaxis(line_slope)

    endIn = is_inside(end_point, one_zone)
    startIn = is_inside(start_point, one_zone)

    color = (0, 255, 0)
    color_code = 0  # 'G'
    color_letter = 'G'
    dir = direction(start_point, end_point)
    if dir > 0:
        # -> use the base_mid_angle_left
        angle = angle_by_slop(line_slope, left_slope)
        use_dir = 1  # 'R'
        use_dir_letter = 'R'
    else:
        # <- use the base_mid_angle_right
        angle = angle_by_slop(line_slope, right_slope)
        use_dir = 0  # 'L'
        use_dir_letter = 'L'
    # already inside
    if (endIn and startIn):
        color = (255, 0, 0)
        color_code = 15  # 'R'1111
        color_letter = 'R'
    elif (endIn and not startIn):
        color = (255, 0, 0)
        color_code = 15  # 'R'1111
        color_letter = 'R'
    elif (not endIn and startIn):
        color_code = 1
        color_letter = 'Y'
        color = (255, 255, 0)
    else:  # both not in
        # most import, not considerate moving up/down direction
        if dir > 0:
            # -> use the base_mid_angle_left
            if (end_point[1] < one_zone.vertices[0][1]):  # above
                # in right side move to -> is safe
                color = (0, 255, 0)
                color_code = 0  # 'G' 0000
                color_letter = 'G'
            else:
                if (end_point[0] >= one_zone.vertices[3][0] and end_point[0] < one_zone.vertices[1][0]):
                    color_code = 1
                    color_letter = 'Y'
                    color = (255, 255, 0)
                elif (end_point[0] < one_zone.vertices[3][0]):
                    color = (0, 255, 0)
                    color_code = 0  # 'G' 0000
                    color_letter = 'G'
                elif (end_point[0] >= one_zone.vertices[1][0]):
                    color = (0, 255, 0)
                    color_code = 0  # 'G' 0000
                    color_letter = 'G'
        else:
            # <-
            if (end_point[1] < one_zone.vertices[0][1]):  # above
                # in right side move to -> is safe
                color = (0, 255, 0)
                color_code = 0  # 'G' 0000
                color_letter = 'G'
            else:
                if (end_point[0] <= one_zone.vertices[2][0] and end_point[0] > one_zone.vertices[1][0]):
                    color_code = 1
                    color_letter = 'Y'
                    color = (255, 255, 0)
                elif (end_point[0] > one_zone.vertices[2][0]):
                    color = (0, 255, 0)
                    color_code = 0  # 'G' 0000
                    color_letter = 'G'
                elif (end_point[0] <= one_zone.vertices[0][0]):
                    color = (0, 255, 0)
                    color_code = 0  # 'G' 0000
                    color_letter = 'G'
    # <- use the base_mid_angle_right
    dist = math.hypot(end_point[0] - start_point[0],
                      end_point[1] - start_point[1])
    return (angle, color_code, color, use_dir, start_point, end_point, color_letter, use_dir_letter, dist)

def Highestrisk(feature_mapping, index):
    for f in feature_mapping:
        kkk = feature_mapping[f]
        prediff = 0
        for j in kkk:
            if (prediff + 2 < len(kkk)):
                post = prediff + 2
                k = prediff
                maxrisk = j[1][1]
                if maxrisk != 15:
                    while k <= post:
                        if kkk[k][1][1] > maxrisk:
                            maxrisk = kkk[k][1][1]
                        k += 1
                sheetdata.append([index, j[0] / fps,
                                  j[1][0],
                                  j[1][1],
                                  j[1][6],
                                  j[1][3],
                                  j[1][7],
                                  j[1][8],
                                  kkk[post][1][0],
                                  kkk[post][1][1],
                                  kkk[post][1][6],
                                  kkk[post][1][3],
                                  kkk[post][1][7],
                                  kkk[post][1][8],
                                  (kkk[post][0] - j[0]) / fps,
                                  maxrisk,
                                  j[1][4][0],
                                  j[1][4][1],
                                  j[1][5][0],
                                  j[1][5][1],
                                  kkk[post][1][4][0],
                                  kkk[post][1][4][1],
                                  kkk[post][1][5][0],
                                  kkk[post][1][5][1]
                                  ])
            prediff += 1


def WriteToExcel(data):
    file_path = 'tracking.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(file_path)


if __name__ == '__main__':
    # model = YOLO('yolov8n_person.pt')  # Load an official Detect model
    # model=YOLO('runs/detect/yolov8n_people_50e18/weights/best.pt',task='detect')  # only people
    model = YOLO("yolov8n.pt")
    res = []
    dir_path = r'./test/'

    if os.path.exists('tracking.xlsx') == False:
        # os.remove('tracking.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            ["videoNo", "TimeStep", "degree", "zone", "zone letter", "L/R", "L/R letter", "dist", "post degree",
             "post zone", "post zone letter", "post L/R", "post L/R letter", "post dist", "gap", "highest risk",
             "startx","starty","endx","endy","post startx","post starty","post endx","post endy"
             ])
        workbook.save('tracking.xlsx')
    df = pd.read_excel('tracking.xlsx')
    # Iterate directory
    for path in os.listdir(dir_path):
        # check if current path is a file
        if os.path.isfile(os.path.join(dir_path, path)):
            res.append(path)
    width = 1280
    height = 720

    fix_left_line = ((int(width * 0.3), int(height * 0.6)), (int(width * 0.2), int(height * 0.85)))
    fix_right_line = ((int(width * 0.7), int(height * 0.6)), (int(width * 0.8), int(height * 0.85)))

    right_mid = find_points_dividing_line(fix_right_line[0], fix_right_line[1], 1 / 2)
    left_mid = find_points_dividing_line(fix_left_line[0], fix_left_line[1], 1 / 2)

    # top_zone = mplPath.Path([, right_mid, left_mid])
    # bottom_zone = mplPath.Path([left_mid, right_mid, fix_right_line[1], fix_left_line[1]])
    one_zone = mplPath.Path([fix_left_line[0], fix_right_line[0], fix_right_line[1], fix_left_line[1]])
    io_flag = 1
    for index in res:
        exists_in_column = index in df["videoNo"].values
        if exists_in_column:
            continue
        sheetdata = []
        sheetdata_cont = []

        VIDEO_Path = dir_path + index
        # Store the track history
        track_history = defaultdict(lambda: [])
        feature_mapping = defaultdict(lambda: [])

        cap = cv2.VideoCapture(VIDEO_Path)
        # width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        # height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        # Get the frames per second (fps) of the video
        # fps = cap.get(cv2.CAP_PROP_FPS)
        fps = 30
        # Calculate the frame interval for every 2 seconds , able to adjust
        frame_interval = 2 * fps
        frame_interval_human = 0.5 * fps

        # width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
        # height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)

        frame_count = 0
        frame_count_human = 0
        # Loop through the video frames
        while cap.isOpened():
            # Read a frame from the video
            success, frame = cap.read()
            if success:
                # fix frame size
                frame = cv2.resize(frame, (width, height))
                # need convert because cv2 is opening in BGR
                # rgb_frame_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                # per 3 sec

                # Run YOLOv8 tracking on the frame, persisting tracks between frames
                results = model.track(frame, persist=True,
                                      tracker='bytetrack.yaml')  # adding tracker cann lower delay =]
                boxes = results[0].boxes.xywh.cpu().numpy()  # this is for the center of the box
                boxes2 = results[0].boxes.xyxy.cpu().numpy()  # this is the actual box
                class_ids = results[0].boxes.cls.cpu().numpy()  # this is the class of each box
                # Visualize the results on the frame
                annotated_frame = frame
                # annotated_frame = results[0].plot()
                if (results[0].boxes.id == None):
                    track_ids = []
                else:
                    track_ids = results[0].boxes.id.int().cpu().tolist()

                # plot the detected lane
                cv2.line(annotated_frame, fix_left_line[0], fix_left_line[1], [255, 255, 255], 5)  # white
                cv2.line(annotated_frame, fix_right_line[0], fix_right_line[1], [0, 255, 0], 5)  # green

                # Plot the tracks, only with the person class with label of tracking, direction
                filtered = [(x, k, y, z) for x, k, y, z in zip(boxes, boxes2, track_ids, class_ids) if z == 0]
                for box, box2, track_id, class_id in filtered:
                    # for box,box2, track_id in zip(boxes,boxes2,track_ids):
                    x, y, w, h = box
                    x1, y1, x2, y2 = map(int, box2[:4])
                    cv2.rectangle(annotated_frame, (x1, y1), (x2, y2), (0, 0, 255), 2)  # red box
                    track = track_history[track_id]
                    track.append((int(x), int(y)))  # x, y center point
                    if len(track) > (frame_interval_human):  # retain 90 tracks for 90 frames
                        track.pop(0)

                    start_point = track[0]
                    end_point = track[len(track) - 1]

                    if (len(track) > 1):
                        # Draw the tracking lines
                        points = np.hstack(track).astype(np.int32).reshape((-1, 1, 2))

                        if frame_count_human % frame_interval_human == 0:
                            # point by ratio

                            # (angle, color_code, color,use_dir,start_point,end_point)
                            moving_feature = features_from_base_point(one_zone, start_point, end_point)
                            feature = feature_mapping[track_id]
                            feature.append((frame_count_human, moving_feature))

                        cv2.line(annotated_frame, (int(start_point[0]), int(start_point[1])),
                                 (int(end_point[0]), int(end_point[1])),
                                 color=(255, 255, 255), thickness=3)

                        # cv2.polylines(annotated_frame, [points], isClosed=False, color=(255, 255, 255),
                        #     thickness=3)

                # Display the annotated frame
                #cv2.imshow("YOLOv8 Tracking", annotated_frame)

                # key = cv2.waitKey(1)  # pauses for 3 seconds before fetching next image
                # Break the loop if 'q' is pressed
                if cv2.waitKey(1) & 0xFF == ord("q"):
                    break
            else:
                # Break the loop if the end of the video is reached
                break
            frame_count += 1
            frame_count_human += 1
            # Release the video capture object and close the display window

        Highestrisk(feature_mapping, index)

        cap.release()
        cv2.destroyAllWindows()

        WriteToExcel(sheetdata)

# TODO: add column highes risk (zone) of the gap in 5 seconds
